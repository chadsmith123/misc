#!/usr/bin/env python
"""
XLSX reader 
Author: Chad Smith 

Outputs xlsx sheet to std out
Requires python 3
"""

import sys
import argparse
from openpyxl import load_workbook

def count(args):
    wb = load_workbook(filename = args.xlsx)
    if args.sheet is None:
        ws = wb[wb.get_sheet_names()[0]]
    else:
        ws = wb[args.sheet]

    for row in ws.iter_rows():
        out=[]
        for cell in row:
            dat=str(cell.value)
            out.append(dat)
        #print(args.delim.join(out))
        print(*out, sep="\t")

    wb.close()
if __name__ == '__main__':
        parser = argparse.ArgumentParser(description='Read xls files')
        parser.add_argument('-sheet', dest='sheet', type=str, help='sheet number')
        #parser.add_argument('-sep', dest='delim', type=str, default=" ", help='Delimiter')
        parser.add_argument('xlsx', help='xls file input')

        if len(sys.argv)==1:
                parser.print_help()
                sys.exit()
        args = parser.parse_args()
        count(args)
